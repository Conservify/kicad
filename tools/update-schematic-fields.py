#!/usr/bin/python

import csv
import argparse
import csv
import copy
import logging
import sys
import shutil
import os
import pprint
import string
import openpyxl as pyxl

from kifield import sch
from kifield import kifield

logger = logging.getLogger('bomtool')
key_fields = [ 'footprint', 'value' ]
fields = [ 'spn1', 'supplier1', 'mfn', 'mfp', 'source', 'notes', 'subsystem', 'critical', 'package' ]
preserved_fields = [ 'footprint', 'value', 'datasheet' ]
pretty_field_names = {
    'spn1': 'SPN1',
    'supplier1': 'Supplier1',
    'mfn': 'MFN',
    'mfp': 'MFP',
    'source': 'Source',
    'notes': 'Notes',
    'subsystem': 'Subsystem',
    'critical': 'Critical',
    'package': 'Package',
    'footprint': 'footprint',
    'value': 'value',
    'price1': 'price1',
    'price100': 'price100',
    'price1000': 'price1000',
    'price5000': 'price5000'
}

class FieldsTablePart:
    def __init__(self, data):
        self.data = data
        self.key = " ".join([data[v] for v in key_fields])

    def first_value(self, keys):
        for key in keys:
            value = self.data.get(pretty_field_names[key], None)
            if value:
                return value
        return ''

    def value(self, key):
        return self.data.get(pretty_field_names[key], '')

    def with_footprint(self, footprint):
        new_data = self.data.copy()
        new_data['footprint'] = footprint
        return FieldsTablePart(new_data)

class FieldsTable:
    def __init__(self, original):
        self.keyed = {}
        self.original = original
        for row in original:
            pi = FieldsTablePart(row)
            self.keyed[pi.key] = pi

class SchematicPart:
    def __init__(self, ref, data):
        self.ref = ref
        self.data = data
        self.key = " ".join([data[v] for v in key_fields])

    def __str__(self):
        return "SchematicPart<%s>" % (self.key,)

    def with_footprint(self, fp):
        new_data = self.data.copy()
        new_data['footprint'] = fp
        return SchematicPart(self.ref, new_data)

    def footprint(self):
        return self.data['footprint']

    def change(self, key, value):
        self.data[key] = value

    def first_value(self, keys):
        for key in keys:
            value = self.data.get(pretty_field_names[key], None)
            if value:
                return value
        return ''

    def value(self, key):
        return self.data.get(pretty_field_names[key], '')

    def remove_fields(self):
        for name in self.data:
            if name not in preserved_fields:
                self.data[name] = ""
        return True

    def to_fields_part(self):
        new_data = self.data.copy()
        new_data['source'] = 'ANY'
        return FieldsTablePart(new_data)

class PartGroup:
    def __init__(self, key, multiplier, parts):
        self.key = key
        self.multiplier = multiplier
        self.parts = parts

    def size(self):
        return "(" + str(len(self.parts)) + "*" + self.multiplier + ")"

    def refs(self):
        values = [part.ref for part in self.parts]
        return ", ".join(values)

    def values(self, key):
        values = [part.value(key) for part in self.parts]
        uniq = set([x for x in values if len(x) > 0])
        return ", ".join(uniq)

class PartGroups:
    def __init__(self, key, groups):
        self.key = key
        self.groups = groups

    def size(self):
        values = [c.size() for c in self.groups]
        return "=(" + (" + ".join(values)) + ")"

    def refs(self):
        values = [c.refs() for c in self.groups]
        return ", ".join(values)

    def values(self, key):
        values = [c.values(key) for c in self.groups]
        uniq = set([x for x in values if len(x) > 0])
        return ", ".join(uniq)

class SchematicTable:
    def __init__(self, original):
        self.original = original
        self.all = []
        self.keyed = {}
        self.groups = None
        for ref, data in original.iteritems():
            part = SchematicPart(ref, data)
            self.keyed[part.key] = self.keyed.get(part.key, []) + [ part ]
            self.all.append(part)

    def sorted(self):
        values = self.all
        values.sort(key=lambda p: p.key)
        return values

    def merge(self, table, multiplier):
        if self.groups is None:
            self.groups = {}

        for key, parts in table.keyed.iteritems():
            self.groups[key] = self.groups.get(key, []) + [ PartGroup(key, multiplier, parts) ]

    def grouped(self):
        rows = [PartGroups(g[0].key, g) for g in self.groups.values()]
        rows.sort(key=lambda pg: pg.key)
        return rows

class UnauthorizedParts:
    def __init__(self):
        self.parts = []

    def add(self, part):
        self.parts.append(part)

    def alternative_keys(self, part):
        fp = part.footprint()
        fp_parts = fp.split(":")
        possible = part.with_footprint('RocketScreamKicadLibrary:' + fp_parts[1])
        return [ possible.key ]

    def resolve(self, source):
        new_rows = []
        for part in self.parts:
            keys = self.alternative_keys(part)
            missing = True
            for key in keys:
                source_part = source.keyed.get(key)
                if source_part is not None:
                    new_rows.append(source_part.with_footprint(part.footprint()))
                    logger.log(logging.INFO, "Found: %s" % (key))
                    missing = False
            if missing:
                new_rows.append(part.to_fields_part())
        for new_part in new_rows:
            source.keyed[new_part.key] = new_part
        return new_rows

class ExcelSubBom:
    def __init__(self, name, quantity_cell):
        self.name = name
        self.quantity = 1
        self.quantity_cell = quantity_cell

class ExcelBom:
    def __init__(self):
        self.wb = pyxl.Workbook()
        self.overview = self.wb.active
        self.overview.title = "overview"
        self.overview.cell(row=1, column=1).value = 'bom'
        self.overview.cell(row=1, column=2).value = 'quantity'
        self.schematics = []
        self.combined_ws = self.wb.create_sheet("combined")

    def prepare(self, files):
        for f in files: self.wb.create_sheet(os.path.basename(f)+ " grouped")
        for f in files: self.wb.create_sheet(os.path.basename(f)+ " itemized")

    def add_schematic_overview(self, filename):
        overview_row = len(self.schematics) + 2
        sub_bom = ExcelSubBom(os.path.basename(filename), "overview!" + 'B' + str(overview_row))
        self.overview.cell(row=overview_row, column=1).value = sub_bom.name
        self.overview.cell(row=overview_row, column=2).value = sub_bom.quantity
        self.schematics.append(sub_bom)
        return sub_bom

    def add_schematic(self, filename, schematic, source_fields):
        sub_bom = self.add_schematic_overview(filename)
        self.individual(sub_bom.name + " itemized", schematic, source_fields)

        grouped = SchematicTable({ })
        grouped.merge(schematic, sub_bom.quantity_cell)
        self.grouped(sub_bom.name + " grouped", grouped, source_fields)

        return sub_bom

    def individual(self, name, schematic, source_fields):
        logger.log(logging.INFO, "Generating '%s' WS" % (name))

        ws = None
        try:
            ws = self.wb[name]
        except KeyError:
            ws = self.wb.create_sheet(name)

        headings = [ 'ref', 'footprint', 'value', 'mfn', 'mfp', 'source', 'critical', 'price1', 'price100', 'price1000', 'price5000' ]
        for c, name in enumerate(headings):
            ws.cell(row=1, column=c + 1).value = name

        for row, part in enumerate(schematic.sorted()):
            source = source_fields.keyed[part.key]
            ws.cell(row=row + 2, column=1).value = part.ref
            ws.cell(row=row + 2, column=2).value = part.value('footprint')
            ws.cell(row=row + 2, column=3).value = part.value('value')
            ws.cell(row=row + 2, column=4).value = part.value('mfn')
            ws.cell(row=row + 2, column=5).value = part.value('mfp')
            ws.cell(row=row + 2, column=6).value = part.value('source')
            ws.cell(row=row + 2, column=7).value = part.value('critical')
            ws.cell(row=row + 2, column=8).value = source.first_value([ 'price1' ])
            ws.cell(row=row + 2, column=9).value = source.first_value([ 'price100', 'price1' ])
            ws.cell(row=row + 2, column=10).value = source.first_value([ 'price1000', 'price100', 'price1' ])
            ws.cell(row=row + 2, column=11).value = source.first_value([ 'price5000', 'price1000', 'price100', 'price1' ])

        return ws

    def grouped(self, name, schematic, source_fields):
        logger.log(logging.INFO, "Generating '%s' WS" % (name))

        ws = None
        try:
            ws = self.wb[name]
        except KeyError:
            ws = self.wb.create_sheet(name)

        headings = [ 'refs', 'footprint', 'value', 'mfn', 'mfp', 'source', 'critical', 'quantity', 'price1', 'price100', 'price1000', 'price5000' ]
        for c, name in enumerate(headings):
            ws.cell(row=1, column=c + 1).value = name

        for row, group in enumerate(schematic.grouped()):
            source = source_fields.keyed[group.key]
            ws.cell(row=row + 2, column=1).value = group.refs()
            ws.cell(row=row + 2, column=2).value = group.values('footprint')
            ws.cell(row=row + 2, column=3).value = group.values('value')
            ws.cell(row=row + 2, column=4).value = group.values('mfn')
            ws.cell(row=row + 2, column=5).value = group.values('mfp')
            ws.cell(row=row + 2, column=6).value = group.values('source')
            ws.cell(row=row + 2, column=7).value = group.values('critical')
            ws.cell(row=row + 2, column=8).value = group.size()
            ws.cell(row=row + 2, column=9).value = source.first_value([ 'price1' ])
            ws.cell(row=row + 2, column=10).value = source.first_value([ 'price100', 'price1' ])
            ws.cell(row=row + 2, column=11).value = source.first_value([ 'price1000', 'price100', 'price1' ])
            ws.cell(row=row + 2, column=12).value = source.first_value([ 'price5000', 'price1000', 'price100', 'price1' ])

        return ws

    def add_combined(self, schematic, source_fields):
        ws = self.grouped("combined", schematic, source_fields)

    def save(self, filename):
        logger.log(logging.INFO, "Saving %s" % (filename))
        self.wb.save(filename)

class ExcelFile:
    def read(self, filename):
        wb = pyxl.load_workbook(filename)
        ws = wb.active
        rows = []
        header = None
        for number, row in enumerate(ws.rows):
            row = [cell.value for cell in row]
            if header is None:
                header = row
                continue

            keyed = {}
            for i, value in enumerate(row):
                if header[i] is not None:
                    keyed[header[i]] = value

            rows.append(keyed)

        self.rows = rows
        self.header = header

        return rows

def change_value(working, filename, from_value, to_value):
    table = SchematicTable(kifield.extract_part_fields_from_sch(filename, recurse=True))

    errors = False
    modified = False
    for key, schematic_parts in table.keyed.iteritems():
        for schematic_part in schematic_parts:
            if schematic_part.value('value') == from_value:
                logger.log(logging.INFO, "Changing: %s: %s" % (os.path.basename(filename), schematic_part))
                schematic_part.change('value', to_value)
                modified = True

    if errors:
        return False

    if modified:
        kifield.insert_part_fields_into_sch(table.original, filename, True, False)

    return True

def update_part_fields(source_part, schematic_part):
    for name in fields:
        schematic_name = pretty_field_names[name]
        if source_part.data.get(name):
            schematic_part.data[schematic_name] = source_part.data[name]
        else:
            schematic_part.data[schematic_name] = " "
    return True

def update_schematic_fields(working, filename, source, unauthorized):
    table = SchematicTable(kifield.extract_part_fields_from_sch(filename, recurse=True))

    errors = False
    modified = False
    for key, schematic_parts in table.keyed.iteritems():
        for schematic_part in schematic_parts:
            source_part = source.keyed.get(schematic_part.key)
            if source_part is None:
                logger.log(logging.ERROR, "No authority: %s: %s" % (os.path.basename(filename), schematic_part))
                unauthorized.add(schematic_part)
                errors = True
                continue
            modified = update_part_fields(source_part, schematic_part) or modified

    if errors:
        return False

    if modified:
        kifield.insert_part_fields_into_sch(table.original, filename, True, False)

    return True

def remove_schematic_fields(working, filename):
    table = SchematicTable(kifield.extract_part_fields_from_sch(filename, recurse=True))

    errors = False
    modified = False
    for key, schematic_parts in table.keyed.iteritems():
        for schematic_part in schematic_parts:
            modified = schematic_part.remove_fields() or modified

    if errors:
        return False

    if modified:
        kifield.insert_part_fields_into_sch(table.original, filename, True, False)

    return True

def read_csv(filename):
    rows = []
    with open(filename) as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            rows.append(row)
    return rows

def update_xlsx_fields(filename, source):
    wb = pyxl.load_workbook(filename)
    ws = wb.active
    header = None
    seen = []
    for number, row in enumerate(ws.rows):
        row = [cell.value for cell in row]
        if header is None:
            header = row
            continue
        keyed = {}
        for i, value in enumerate(row):
            if header[i] is not None:
                keyed[header[i]] = value

        key = " ".join([keyed[v] for v in key_fields])
        f = source.keyed.get(key)
        if f:
            for name in [ 'mfn', 'mfp' ]:
                column = header.index(name)
                ws.cell(row=number + 1, column = column + 1).value = f.data[name]
            seen.append(key)
        else:
            logger.log(logging.ERROR, "Missing %s", key)

    bottom = ws.max_row

    for part in source.keyed.values():
        if part.key not in seen:
            pprint.pprint([ 'Adding to XLSX: ', bottom, part.key ])
            for index, name in enumerate(header):
                value = ""
                if name in part.data:
                    value = part.data[name]
                ws.cell(row = bottom, column = index + 1).value = value
            bottom += 1

    updated_fn = os.path.join(os.path.dirname(filename), "updated-" + os.path.basename(filename))
    wb.save(updated_fn)

def configure_logging():
    logging.basicConfig(format='%(asctime)-15s %(levelname)s %(message)s')
    if False: logging.getLogger('kifield').setLevel(logging.DEBUG - 0)
    logger.setLevel(logging.DEBUG)

def main():
    configure_logging()

    parser = argparse.ArgumentParser(description='')
    parser.add_argument('--remove', help='', action="store_true")
    parser.add_argument('--update', help='', action="store_true")
    parser.add_argument('--bom', help='', action="store_true")
    parser.add_argument('--from-value', help='', action="store")
    parser.add_argument('--to-value', help='', action="store")
    parser.add_argument('args', nargs='*')

    args = parser.parse_args()
    working = os.getcwd()

    processed_args = []
    for arg in args.args:
        if os.path.isfile(arg):
            processed_args.append(os.path.abspath(arg))

    errors = False

    logger.log(logging.INFO, "Opening %s" % ("authority.xlsx"))

    authority = ExcelFile()
    authority_fn = os.path.abspath("authority.xlsx")
    source_fields = FieldsTable(authority.read(authority_fn))
    unauthorized = UnauthorizedParts()

    if args.remove:
        for child_filename in processed_args:
            if os.path.isfile(child_filename):
                os.chdir(os.path.dirname(child_filename))

                logger.log(logging.INFO, "Removing fields %s" % (child_filename))
                errors = not remove_schematic_fields(working, child_filename) or errors

        if errors:
            return

    if args.update:
        for child_filename in processed_args:
            if os.path.isfile(child_filename):
                os.chdir(os.path.dirname(child_filename))

                logger.log(logging.INFO, "Updating fields %s" % (child_filename))
                errors = not update_schematic_fields(working, child_filename, source_fields, unauthorized) or errors

        rows = unauthorized.resolve(source_fields)
        update_xlsx_fields(authority_fn, source_fields)

        if errors:
            return

    if args.from_value and args.to_value:
        logger.log(logging.INFO, "Changing value '%s' to '%s'" % (args.from_value, args.to_value))
        for child_filename in processed_args:
            logger.log(logging.INFO, "Processing %s" % (child_filename))
            os.chdir(os.path.dirname(child_filename))

            errors = not change_value(working, child_filename, args.from_value, args.to_value) or errors

    if args.bom:
        combined = SchematicTable({ })

        super_bom = ExcelBom()
        super_bom.prepare(processed_args)

        for arg in processed_args:
            child_filename = arg
            os.chdir(os.path.dirname(child_filename))

            logger.log(logging.INFO, "Processing %s" % (child_filename))

            if not args.update:
                logger.log(logging.INFO, "Updating fields %s" % (child_filename))
                errors = not update_schematic_fields(working, child_filename, source_fields, unauthorized) or errors

            schematic = SchematicTable(kifield.extract_part_fields_from_sch(child_filename, recurse=True))
            sub_bom = super_bom.add_schematic(child_filename, schematic, source_fields)
            combined.merge(schematic, sub_bom.quantity_cell)

        if errors:
            return

        super_bom.add_combined(combined, source_fields)
        super_bom.save(os.path.join(working, "super.xlsx"))

if __name__ == "__main__":
    main()
